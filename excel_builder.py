import logging
import os
import shutil
import tempfile
from datetime import datetime

import openpyxl

logger = logging.getLogger(__name__)

TEMPLATE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "template.xlsx")

# Product code to row mapping in the template
PRODUCT_ROWS = {
    "BOSH0077": 46,
    "BOSH0079": 47,
    "BOSH0081": 48,
    "BOSH0080": 49,
    "BOSH0090": 50,
    "BOSH0082": 51,
    "BOSH0089": 52,
}


def build_excel(delivery_date: str, quantities: dict[str, int]) -> str:
    """Populate the Excel template with PO data.

    Args:
        delivery_date: Date string in YYYY-MM-DD format.
        quantities: Dict of product_code -> quantity.

    Returns:
        Path to the generated Excel file.
    """
    dt = datetime.strptime(delivery_date, "%Y-%m-%d")

    # Copy template to temp file
    filename = f"RPA {dt.strftime('%d_%m')}.xlsx"
    tmp_dir = tempfile.mkdtemp(prefix="po_excel_")
    excel_path = os.path.join(tmp_dir, filename)
    shutil.copy2(TEMPLATE_PATH, excel_path)

    wb = openpyxl.load_workbook(excel_path)
    ws = wb["Sheet1"]

    # Set delivery date (C41)
    ws["C41"] = dt
    logger.info("Set C41 (Delivery Date) = %s", delivery_date)

    # Set product quantities (H46:H52)
    for product_code, row in PRODUCT_ROWS.items():
        qty = quantities.get(product_code, 0)
        ws[f"H{row}"] = qty
        logger.info("Set H%d (%s) = %d", row, product_code, qty)

    wb.save(excel_path)
    logger.info("Excel saved: %s", excel_path)
    return excel_path
